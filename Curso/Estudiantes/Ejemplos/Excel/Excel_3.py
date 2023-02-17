import re
import time
import random
import pytest
import openpyxl
from playwright.sync_api import Page, expect,Playwright, sync_playwright
from funciones_excel import Funciones_Globales

#pytest Sesion2.py -s -v
#pytest Excel_2.py -s -v --browser-channel=chrome -n 4
#v-17
#playwright codegen   https://www.saucedemo.com/
#https://www.mockaroo.com/

#Variables globales
tiempo=0.9
ruta="Imagenes/Upload/"
pdf1="C:/Users/Usuario1/Documents/VIDEOS_UDEMY/PLAYWRIGHT/Curso/Practicas/Ejemplos/Pdf/Documento1.pdf"
ruta="C:/Users/Rodrigo/Documents/VIDEOS_UDEMY_ES/PLAYWRIGHT/Curso/Estudiantes/Ejemplos/Excel/datos.xlsx"
#registros=6

##Leyendo el archivo de Excel
archivo=openpyxl.load_workbook(ruta)
def Numero_filas(hoja):
    ac=archivo[hoja]
    return  ac.max_row
  
def Dato_columna(hoja,fila,col):
    ac=archivo[hoja]
    col=ac.cell(int(fila),int(col))
    return  col.value

print(Numero_filas("Hoja1"))
print(Dato_columna("Hoja1",1,2))

def test_excel1(set_up_excel)-> None:
    page=set_up_excel    
    F=Funciones_Globales(page)
    F.Validar_titulo_pagina("Datos Personales | TestingQaRvn")
    F.Scroll_xy(0,600)
    print("Cargando excel")
    F.Esperar(3)


    for n in range(2,10):
      nombre=  Dato_columna("Hoja1",n,1)
      print(nombre)
      ap=  Dato_columna("Hoja1",n,2)
      print(ap)
      email=  Dato_columna("Hoja1",n,3)
      print(email)
      telefono=  Dato_columna("Hoja1",n,4)
      print(telefono)
      direccion=  Dato_columna("Hoja1",n,5)
      print(direccion)
      
      F.Texto("//input[contains(@id,'wsf-1-field-21')]",nombre)
      F.Texto("//input[contains(@id,'wsf-1-field-22')]",ap)
      F.Texto("//input[contains(@id,'wsf-1-field-23')]",email)
      F.Texto("//input[contains(@id,'wsf-1-field-24')]",str(telefono))
      F.Texto("//textarea[contains(@id,'wsf-1-field-28')]",direccion)
      F.Click("//button[contains(@id,'wsf-1-field-27')]")
      F.Esperar(1)
      page.goto("https://testingqarvn.com.es/datos-personales/")
      F.Esperar(1)
      

def test_excel2(set_up_excel)-> None:
    page=set_up_excel    
    F=Funciones_Globales(page)
    F.Validar_titulo_pagina("Datos Personales | TestingQaRvn")
    F.Scroll_xy(0,600)
    print("Cargando excel")
    F.Esperar(3)


    for n in range(11,22):
      nombre=  Dato_columna("Hoja1",n,1)
      print(nombre)
      ap=  Dato_columna("Hoja1",n,2)
      print(ap)
      email=  Dato_columna("Hoja1",n,3)
      print(email)
      telefono=  Dato_columna("Hoja1",n,4)
      print(telefono)
      direccion=  Dato_columna("Hoja1",n,5)
      print(direccion)
      
      F.Texto("//input[contains(@id,'wsf-1-field-21')]",nombre)
      F.Texto("//input[contains(@id,'wsf-1-field-22')]",ap)
      F.Texto("//input[contains(@id,'wsf-1-field-23')]",email)
      F.Texto("//input[contains(@id,'wsf-1-field-24')]",str(telefono))
      F.Texto("//textarea[contains(@id,'wsf-1-field-28')]",direccion)
      F.Click("//button[contains(@id,'wsf-1-field-27')]")
      F.Esperar(1)
      page.goto("https://testingqarvn.com.es/datos-personales/")
      F.Esperar(1)
      
def test_excel3(set_up_excel)-> None:
    page=set_up_excel    
    F=Funciones_Globales(page)
    F.Validar_titulo_pagina("Datos Personales | TestingQaRvn")
    F.Scroll_xy(0,600)
    print("Cargando excel")
    F.Esperar(3)


    for n in range(23,44):
      nombre=  Dato_columna("Hoja1",n,1)
      print(nombre)
      ap=  Dato_columna("Hoja1",n,2)
      print(ap)
      email=  Dato_columna("Hoja1",n,3)
      print(email)
      telefono=  Dato_columna("Hoja1",n,4)
      print(telefono)
      direccion=  Dato_columna("Hoja1",n,5)
      print(direccion)
      
      F.Texto("//input[contains(@id,'wsf-1-field-21')]",nombre)
      F.Texto("//input[contains(@id,'wsf-1-field-22')]",ap)
      F.Texto("//input[contains(@id,'wsf-1-field-23')]",email)
      F.Texto("//input[contains(@id,'wsf-1-field-24')]",str(telefono))
      F.Texto("//textarea[contains(@id,'wsf-1-field-28')]",direccion)
      F.Click("//button[contains(@id,'wsf-1-field-27')]")
      F.Esperar(1)
      page.goto("https://testingqarvn.com.es/datos-personales/")
      F.Esperar(1)
      
def test_excel4(set_up_excel)-> None:
    page=set_up_excel    
    F=Funciones_Globales(page)
    F.Validar_titulo_pagina("Datos Personales | TestingQaRvn")
    F.Scroll_xy(0,600)
    print("Cargando excel")
    F.Esperar(3)


    for n in range(44,55):
      nombre=  Dato_columna("Hoja1",n,1)
      print(nombre)
      ap=  Dato_columna("Hoja1",n,2)
      print(ap)
      email=  Dato_columna("Hoja1",n,3)
      print(email)
      telefono=  Dato_columna("Hoja1",n,4)
      print(telefono)
      direccion=  Dato_columna("Hoja1",n,5)
      print(direccion)
      
      F.Texto("//input[contains(@id,'wsf-1-field-21')]",nombre)
      F.Texto("//input[contains(@id,'wsf-1-field-22')]",ap)
      F.Texto("//input[contains(@id,'wsf-1-field-23')]",email)
      F.Texto("//input[contains(@id,'wsf-1-field-24')]",str(telefono))
      F.Texto("//textarea[contains(@id,'wsf-1-field-28')]",direccion)
      F.Click("//button[contains(@id,'wsf-1-field-27')]")
      F.Esperar(1)
      page.goto("https://testingqarvn.com.es/datos-personales/")
      F.Esperar(1)
      
def test_excel5(set_up_excel)-> None:
    page=set_up_excel    
    F=Funciones_Globales(page)
    F.Validar_titulo_pagina("Datos Personales | TestingQaRvn")
    F.Scroll_xy(0,600)
    print("Cargando excel")
    F.Esperar(3)


    for n in range(56,67):
      nombre=  Dato_columna("Hoja1",n,1)
      print(nombre)
      ap=  Dato_columna("Hoja1",n,2)
      print(ap)
      email=  Dato_columna("Hoja1",n,3)
      print(email)
      telefono=  Dato_columna("Hoja1",n,4)
      print(telefono)
      direccion=  Dato_columna("Hoja1",n,5)
      print(direccion)
      
      F.Texto("//input[contains(@id,'wsf-1-field-21')]",nombre)
      F.Texto("//input[contains(@id,'wsf-1-field-22')]",ap)
      F.Texto("//input[contains(@id,'wsf-1-field-23')]",email)
      F.Texto("//input[contains(@id,'wsf-1-field-24')]",str(telefono))
      F.Texto("//textarea[contains(@id,'wsf-1-field-28')]",direccion)
      F.Click("//button[contains(@id,'wsf-1-field-27')]")
      F.Esperar(1)
      page.goto("https://testingqarvn.com.es/datos-personales/")
      F.Esperar(1)
      